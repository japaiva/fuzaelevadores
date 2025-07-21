# producao/views/relatorios_produtos.py - NOVO: Relatório completo de produtos

"""
Relatório completo de produtos com filtros e exportação Excel
Portal de Produção - Sistema Elevadores FUZA
"""

import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q
from django.utils import timezone
from io import BytesIO

from core.models import Produto, GrupoProduto, SubgrupoProduto

logger = logging.getLogger(__name__)

@login_required
def relatorio_produtos_completo(request):
    """
    Relatório completo de produtos com filtros avançados
    """
    # Base query
    produtos_query = Produto.objects.select_related(
        'grupo', 'subgrupo', 'fornecedor_principal'
    ).order_by('codigo')
    
    # Aplicar filtros
    filtros_aplicados = {}
    
    # Filtro por tipo de produto
    tipo_produto = request.GET.get('tipo_produto')
    if tipo_produto in ['MP', 'PI', 'PA']:
        produtos_query = produtos_query.filter(tipo=tipo_produto)
        filtros_aplicados['tipo_produto'] = tipo_produto
    
    # Filtro por grupo
    grupo_id = request.GET.get('grupo')
    if grupo_id and grupo_id.isdigit():
        produtos_query = produtos_query.filter(grupo_id=grupo_id)
        filtros_aplicados['grupo'] = grupo_id
    
    # Filtro por subgrupo
    subgrupo_id = request.GET.get('subgrupo')
    if subgrupo_id and subgrupo_id.isdigit():
        produtos_query = produtos_query.filter(subgrupo_id=subgrupo_id)
        filtros_aplicados['subgrupo'] = subgrupo_id
    
    # Filtro por tipo PI
    tipo_pi = request.GET.get('tipo_pi')
    if tipo_pi and tipo_pi in dict(Produto.TIPO_PI_CHOICES):
        produtos_query = produtos_query.filter(tipo_pi=tipo_pi)
        filtros_aplicados['tipo_pi'] = tipo_pi
    
    # Filtro por status ativo
    status_ativo = request.GET.get('status_ativo')
    if status_ativo == 'sim':
        produtos_query = produtos_query.filter(status='ATIVO')
        filtros_aplicados['status_ativo'] = 'sim'
    elif status_ativo == 'nao':
        produtos_query = produtos_query.filter(status='INATIVO')
        filtros_aplicados['status_ativo'] = 'nao'
    
    # Filtro por utilizado
    utilizado = request.GET.get('utilizado')
    if utilizado == 'sim':
        produtos_query = produtos_query.filter(utilizado=True)
        filtros_aplicados['utilizado'] = 'sim'
    elif utilizado == 'nao':
        produtos_query = produtos_query.filter(utilizado=False)
        filtros_aplicados['utilizado'] = 'nao'
    
    # Busca por texto
    query = request.GET.get('q')
    if query:
        produtos_query = produtos_query.filter(
            Q(codigo__icontains=query) |
            Q(nome__icontains=query) |
            Q(descricao__icontains=query)
        )
        filtros_aplicados['query'] = query
    
    # Executar query
    produtos = produtos_query.all()
    
    # Verificar se é solicitação de exportação
    if request.GET.get('export') == 'excel':
        return exportar_produtos_excel(request, produtos, filtros_aplicados)
    
    # Dados para os filtros
    context = {
        'produtos': produtos,
        'total_produtos': produtos.count(),
        'filtros': filtros_aplicados,
        
        # Para os selects de filtro
        'grupos': GrupoProduto.objects.filter(ativo=True).order_by('codigo'),
        'subgrupos': SubgrupoProduto.objects.filter(ativo=True).select_related('grupo').order_by('grupo__codigo', 'codigo'),
        'tipo_pi_choices': Produto.TIPO_PI_CHOICES,
        
        # Valores dos filtros para manter selecionados
        'tipo_produto_filtro': tipo_produto,
        'grupo_filtro': grupo_id,
        'subgrupo_filtro': subgrupo_id,
        'tipo_pi_filtro': tipo_pi,
        'status_ativo_filtro': status_ativo,
        'utilizado_filtro': utilizado,
        'query_filtro': query,
    }
    
    return render(request, 'producao/relatorios/produtos_completo.html', context)


def exportar_produtos_excel(request, produtos, filtros_aplicados):
    """
    Exporta relatório de produtos para Excel
    """
    logger.info(f'📊 Exportando {produtos.count()} produtos para Excel')
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório Produtos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalho do relatório
    ws.merge_cells('A1:M1')
    ws['A1'] = "RELATÓRIO COMPLETO DE PRODUTOS - SISTEMA FUZA"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Data e filtros
    linha_atual = 2
    ws[f'A{linha_atual}'] = f"Data: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    linha_atual += 1
    
    if filtros_aplicados:
        ws[f'A{linha_atual}'] = "Filtros aplicados:"
        linha_atual += 1
        for filtro, valor in filtros_aplicados.items():
            if filtro == 'tipo_produto':
                ws[f'A{linha_atual}'] = f"• Tipo: {dict(Produto.TIPO_CHOICES).get(valor, valor)}"
            elif filtro == 'grupo':
                try:
                    grupo = GrupoProduto.objects.get(id=valor)
                    ws[f'A{linha_atual}'] = f"• Grupo: {grupo.codigo} - {grupo.nome}"
                except:
                    ws[f'A{linha_atual}'] = f"• Grupo: {valor}"
            elif filtro == 'subgrupo':
                try:
                    subgrupo = SubgrupoProduto.objects.get(id=valor)
                    ws[f'A{linha_atual}'] = f"• Subgrupo: {subgrupo.codigo_completo} - {subgrupo.nome}"
                except:
                    ws[f'A{linha_atual}'] = f"• Subgrupo: {valor}"
            elif filtro == 'tipo_pi':
                ws[f'A{linha_atual}'] = f"• Tipo PI: {dict(Produto.TIPO_PI_CHOICES).get(valor, valor)}"
            elif filtro == 'status_ativo':
                ws[f'A{linha_atual}'] = f"• Status: {'Ativo' if valor == 'sim' else 'Inativo'}"
            elif filtro == 'utilizado':
                ws[f'A{linha_atual}'] = f"• Utilizado: {'Sim' if valor == 'sim' else 'Não'}"
            elif filtro == 'query':
                ws[f'A{linha_atual}'] = f"• Busca: {valor}"
            
            linha_atual += 1
    
    linha_atual += 1
    
    # Cabeçalhos das colunas
    headers = [
        'Código',
        'Nome',
        'Tipo',
        'Grupo',
        'Subgrupo',
        'Tipo PI',
        'Custo Material',
        'Custo Serviço',
        'Custo Total',
        'Utilizado',
        'Status'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=linha_atual, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_thin
    
    linha_atual += 1
    
    # Dados dos produtos
    for produto in produtos:
        row_data = [
            produto.codigo or '',
            produto.nome or '',
            produto.get_tipo_display() or '',
            f"{produto.grupo.codigo} - {produto.grupo.nome}" if produto.grupo else '',
            f"{produto.subgrupo.codigo_completo} - {produto.subgrupo.nome}" if produto.subgrupo else '',
            produto.get_tipo_pi_display() if produto.tipo_pi else '',
            float(produto.custo_material or 0),
            float(produto.custo_servico or 0),
            float(produto.custo_total if hasattr(produto, 'custo_total') else 0),
            'Sim' if produto.utilizado else 'Não',
            'Ativo' if produto.status == 'ATIVO' else 'Inativo'
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=linha_atual, column=col)
            cell.value = value
            cell.border = border_thin
            
            # Formatação para valores monetários
            if col in [8, 9, 10]:  # Colunas de custo
                cell.number_format = 'R$ #,##0.00'
            elif col == 11:  # Coluna de estoque
                cell.number_format = '#,##0.00'
        
        linha_atual += 1
    
    # Totais
    linha_atual += 1
    ws[f'A{linha_atual}'] = f"Total de produtos: {produtos.count()}"
    ws[f'A{linha_atual}'].font = Font(bold=True)
    
    # Ajustar largura das colunas
    column_widths = {
        'A': 15,  # Código
        'B': 40,  # Nome
        'D': 12,  # Tipo
        'E': 30,  # Grupo
        'F': 30,  # Subgrupo
        'G': 20,  # Tipo PI
        'H': 15,  # Custo Material
        'I': 15,  # Custo Serviço
        'J': 15,  # Custo Total
        'L': 10,  # Utilizado
        'M': 10,  # Status
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Salvar em memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Criar response
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'relatorio_produtos_{timestamp}.xlsx'
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    logger.info(f'✅ Excel exportado: {filename}')
    return response


@login_required
def api_subgrupos_por_grupo_relatorio(request):
    """
    API para carregar subgrupos baseado no grupo para o relatório
    """
    from django.http import JsonResponse
    
    grupo_id = request.GET.get('grupo_id')
    
    if not grupo_id:
        return JsonResponse({'success': True, 'subgrupos': []})
    
    try:
        subgrupos = SubgrupoProduto.objects.filter(
            grupo_id=grupo_id, 
            ativo=True
        ).select_related('grupo').order_by('codigo')
        
        subgrupos_data = []
        for subgrupo in subgrupos:
            subgrupos_data.append({
                'id': subgrupo.id,
                'codigo': subgrupo.codigo,
                'nome': subgrupo.nome,
                'codigo_completo': subgrupo.codigo_completo,
            })
        
        return JsonResponse({
            'success': True,
            'subgrupos': subgrupos_data
        })
        
    except Exception as e:
        logger.error(f'Erro ao buscar subgrupos para relatório: {str(e)}')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)